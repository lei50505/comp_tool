#! /usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import *
from openpyxl.styles import *
from openpyxl.styles.numbers import *

from util import to_str, to_float

import os

class Cell():
    '''doc'''
    def __init__(self, cell):
        self.cell = cell

    def get_val(self):
        '''doc'''
        return self.cell.value

    def get_float_val(self):
        '''doc'''
        val = self.get_val()
        return to_float(val)

    def get_str_val(self):
        '''doc'''
        val = self.get_val()
        return to_str(val)

    def set_val(self, val):
        '''doc'''
        self.cell.value = val

    def setNumberFormatText(self):
        '''doc'''
        self.cell.number_format = numbers.FORMAT_TEXT

    def setFillRed(self):
        '''doc'''
        self.cell.fill = PatternFill(fill_type = "solid",\
                            start_color="FFCCFF",end_color="FFCCFF")

    def setFillBlue(self):
        '''doc'''
        self.cell.fill = PatternFill(fill_type = "solid", \
                            start_color="CCFFFF", end_color="CCFFFF")
        
    def setBorderThin(self):
        '''doc'''
        thinSide = Side(border_style="thin", color="000000")
        thinBorder = Border(top=thinSide, left=thinSide, \
                            right=thinSide, bottom=thinSide)
        self.cell.border = thinBorder

class Sheet():
    def __init__(self, sheet):
        self.sheet = sheet
        
        self.numColIndex = None
        
        self.numValDict = None
        self.numValSet = None
        self.numValList = None
        
        self.diffNumRows = None
        
        self.numRowDict = None
        self.numRowList = None

        self.copyRowCount = 0

    def cell(self, row, col):
        '''doc'''
        sheetCell = self.sheet.cell(row = row, column = col)
        cell = Cell(sheetCell)
        return cell

    def getMaxCol(self):
        '''doc'''
        return self.sheet.max_column

    def getMaxRow(self):
        '''doc'''
        return self.sheet.max_row

    def initNumColIndex(self):
        '''doc'''
        maxCol = self.getMaxCol()
        maxRow = self.getMaxRow()
        
        numColCount = 0
        numColIndex = 0
        
        for colIndex in range(1, maxCol + 1):
            numCellCount = 0
            isStrCell = False
            for rowIndex in range(1, maxRow + 1):
                cell = self.cell(rowIndex, colIndex)
                floatVal = cell.get_float_val()

                if isinstance(floatVal, float):
                    numCellCount += 1
                    continue

                strVal = cell.get_str_val()
                
                if isinstance(strVal, str):
                    isStrCell = True
                    break

                
                
            if numCellCount >= 1 and not isStrCell:
                numColCount += 1
                numColIndex = colIndex

        if numColCount == 1:
            self.numColIndex = numColIndex

        if numColCount == 0:
            raise Exception("没有数字列")

        if numColCount > 1:
            raise Exception("有%d列是数字" % numColCount)

        

    def initNumColDict(self):
        '''doc'''
        
        if self.numColIndex is None:
            raise Exception("请先初始化numColIndex")
        
        self.numValDict = {}
        self.numValSet = set()
        self.numValList = []
        
        self.numRowDict = {}
        self.numRowList = []

        numColIndex = self.numColIndex
        maxRow = self.getMaxRow()
        maxCol = self.getMaxCol()
        
        for rowIndex in range(1, maxRow + 1):
            cell = self.cell(rowIndex, self.numColIndex)
            floatVal = cell.get_float_val()
            if isinstance(floatVal,float):
                self.numValSet.add(floatVal)
                self.numValList.append(floatVal)
                self.numRowDict[rowIndex] = floatVal
                self.numRowList.append(rowIndex)
                
                dictVal = self.numValDict.get(floatVal)
                if dictVal is None:
                    self.numValDict[floatVal] = 1
                    continue
                
                self.numValDict[floatVal] = dictVal + 1

                
    def initDiffNumRows(self):
        '''doc'''
        self.diffNumRows = []
        for numRow in self.numRowList:
            val = self.numRowDict.get(numRow)
            count = self.numValList.count(val)
            if count == 1:
                self.diffNumRows.append(numRow)

    def getRowListByVal(self, val):
        '''doc'''
        val = to_float(val)
        ret = []
        for numRow in self.numRowList:
            v = self.numRowDict.get(numRow)
            if v == val:
                ret.append(numRow)
        return ret

    def copyRowFromSheet(self, srcSheet, rowIndex, color):
        '''doc'''

        copyRowCount = self.copyRowCount + 1
        srcSheetMaxCol = srcSheet.getMaxCol()
        
        for srcSheetColIndex in range(1, srcSheetMaxCol + 1):
            srcCell = srcSheet.cell(rowIndex, srcSheetColIndex)
            tarCell = self.cell(copyRowCount, srcSheetColIndex)
            tarCell.set_val(srcCell.get_val())

            tarCell.setBorderThin()
            
            tarCell.setNumberFormatText()

            if "red" == color:
                tarCell.setFillRed()
            elif "blue" == color:
                tarCell.setFillBlue()
            
        self.copyRowCount += 1
        
  

class Book():
    def __init__(self, book):
        self.book=book

    def active(self):
        '''doc'''
        activeSheet = self.book.active
        return Sheet(activeSheet)
        
    def sheet(self,sheetName):
        '''doc'''
       
        bookSheet = self.book[sheetName]
        sheet = Sheet(bookSheet)
        return sheet


    def hasSheet(self, *sheetNames):
        '''doc'''

        if len(sheetNames) == 0:
            return True
        
        bookSheetNames = self.book.get_sheet_names()
        for sheetName in sheetNames:
            if not isinstance(sheetName,str):
                return False
            if sheetName not in bookSheetNames:
                return False
        return True

    def save(self, path):
        '''doc'''
        self.book.save(path)

    def close(self):
        '''doc'''
        if self.book is not None:
            self.book.close()
        

def create_book():
    '''doc'''
    work_book =  Workbook(write_only=False)
    book = Book(work_book)
    return book

def load_book(file_path):
    '''doc'''

    work_book = load_workbook(file_path, read_only = True, keep_vba = False, \
                    data_only = True, guess_types = False, keep_links = False)
    book = Book(work_book)
    return book

